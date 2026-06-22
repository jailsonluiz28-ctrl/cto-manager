from django.http import HttpResponse
from django.shortcuts import render, get_object_or_404, redirect
from django.db.models import Q

from .models import CTO
from .forms import CTOForm
from clientes.models import (
    Cliente,
    HistoricoMovimentacao
)

import openpyxl

from reportlab.lib.pagesizes import letter
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
)
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet

from datetime import datetime

from django.contrib.auth.decorators import login_required
from accounts.permissoes import (
    operador_required
)


@login_required
def lista_ctos(request):

    busca = request.GET.get('busca', '').strip()

    ctos = CTO.objects.all()

    if busca:
        ctos = ctos.filter(
            Q(nome__icontains=busca) |
            Q(rua__icontains=busca) |
            Q(tipo__icontains=busca)
        )

    ranking = []

    for cto in ctos:

        if cto.portas_total > 0:
            percentual = round(
                (cto.portas_ocupadas / cto.portas_total) * 100,
                1
            )
        else:
            percentual = 0

        ranking.append({
            'cto': cto,
            'percentual': percentual,
        })

    ranking.sort(
        key=lambda item: item['percentual'],
        reverse=True
    )

    total_portas = 0
    total_ocupadas = 0

    for cto in ctos:
        total_portas += cto.portas_total
        total_ocupadas += cto.portas_ocupadas

    total_livres = total_portas - total_ocupadas

    if total_portas > 0:
        taxa_ocupacao = round(
            (total_ocupadas / total_portas) * 100,
            1
        )
    else:
        taxa_ocupacao = 0

    return render(
        request,
        'ctos/lista_ctos.html',
        {
            'ctos': ctos,
            'busca': busca,
            'ranking': ranking,

            'total_portas': total_portas,
            'total_ocupadas': total_ocupadas,
            'total_livres': total_livres,
            'taxa_ocupacao': taxa_ocupacao,

            'eh_admin': request.user.groups.filter(
                name='Administrador'
            ).exists(),

            'eh_operador': request.user.groups.filter(
                name='Operador'
            ).exists(),

            'eh_tecnico': request.user.groups.filter(
                name='Consulta do Técnico'
            ).exists(),
        }
    )


@login_required
def detalhe_cto(request, cto_id):

    cto = get_object_or_404(
        CTO,
        id=cto_id
    )

    clientes = Cliente.objects.filter(
        cto=cto
    ).order_by('porta')

    portas_livres = (
        cto.portas_total -
        cto.portas_ocupadas
    )

    mapa_portas = []

    for numero in range(
        1,
        cto.portas_total + 1
    ):

        cliente_porta = clientes.filter(
            porta=numero
        ).first()

        if cliente_porta:

            mapa_portas.append({
                'numero': numero,
                'status': 'ocupada',
                'cliente': cliente_porta.nome,
            })

        else:

            mapa_portas.append({
                'numero': numero,
                'status': 'livre',
                'cliente': '',
            })

    return render(
        request,
        'ctos/detalhe_cto.html',
        {
            'cto': cto,
            'clientes': clientes,
            'portas_livres': portas_livres,
            'mapa_portas': mapa_portas,
        }
    )


@login_required
def nova_cto(request):

    if request.method == 'POST':

        form = CTOForm(request.POST)

        if form.is_valid():

            cto = form.save()

            HistoricoMovimentacao.objects.create(
                usuario=request.user.username,
                cliente_nome="-",
                cto_nome=cto.nome,
                porta=0,
                acao="CTO CRIADA",
                observacao=f"Rua: {cto.rua}"
            )

            return redirect('/ctos/')

    else:

        form = CTOForm()

    return render(
        request,
        'ctos/nova_cto.html',
        {
            'form': form
        }
    )


@login_required
@operador_required
def editar_cto(
    request,
    cto_id
):

    cto = CTO.objects.get(
        id=cto_id
    )

    if request.method == 'POST':

        form = CTOForm(
            request.POST,
            instance=cto
        )

        if form.is_valid():

            cto_editada = form.save()

            HistoricoMovimentacao.objects.create(
                usuario=request.user.username,
                cliente_nome="-",
                cto_nome=cto_editada.nome,
                porta=0,
                acao="CTO EDITADA",
                observacao=f"Rua: {cto_editada.rua}"
            )

            return redirect(
                '/ctos/'
            )

    else:

        form = CTOForm(
            instance=cto
        )

    return render(
        request,
        'ctos/editar_cto.html',
        {
            'form': form,
            'cto': cto,
        }
    )


@login_required
def exportar_ctos_excel(request):

    workbook = openpyxl.Workbook()

    ctos = CTO.objects.all()

    total_ctos = ctos.count()

    total_portas = sum(
        cto.portas_total for cto in ctos
    )

    total_ocupadas = sum(
        cto.portas_ocupadas for cto in ctos
    )

    total_livres = (
        total_portas - total_ocupadas
    )

    if total_portas > 0:
        taxa_ocupacao = round(
            (total_ocupadas / total_portas) * 100,
            1
        )
    else:
        taxa_ocupacao = 0

    # =========================
    # ABA 1 - RESUMO EXECUTIVO
    # =========================

    resumo = workbook.active
    resumo.title = 'Resumo Executivo'

    resumo.append(['Indicador', 'Valor'])

    resumo.append(['Total de CTOs', total_ctos])
    resumo.append(['Total de Portas', total_portas])
    resumo.append(['Portas Ocupadas', total_ocupadas])
    resumo.append(['Portas Livres', total_livres])
    resumo.append(['Taxa de Ocupação (%)', taxa_ocupacao])

    # =========================
    # ABA 2 - RANKING
    # =========================

    ranking_ws = workbook.create_sheet(
        title='Ranking'
    )

    ranking_ws.append([
        'CTO',
        'Ocupadas',
        'Totais',
        'Percentual'
    ])

    ranking = []

    for cto in ctos:

        if cto.portas_total > 0:

            percentual = round(
                (
                    cto.portas_ocupadas /
                    cto.portas_total
                ) * 100,
                1
            )

        else:

            percentual = 0

        ranking.append([
            cto.nome,
            cto.portas_ocupadas,
            cto.portas_total,
            percentual
        ])

    ranking.sort(
        key=lambda linha: linha[3],
        reverse=True
    )

    for linha in ranking:
        ranking_ws.append(linha)

    # =========================
    # ABA 3 - CTOs
    # =========================

    detalhes = workbook.create_sheet(
        title='CTOs'
    )

    detalhes.append([
        'CTO',
        'Tipo',
        'Rua',
        'Portas Totais',
        'Portas Ocupadas',
        'Portas Livres',
        'Ativa'
    ])

    for cto in ctos:

        detalhes.append([
            cto.nome,
            cto.tipo,
            cto.rua,
            cto.portas_total,
            cto.portas_ocupadas,
            cto.portas_total - cto.portas_ocupadas,
            'Sim' if cto.ativa else 'Não',
        ])

    # Ajuste automático das colunas

    for aba in workbook.worksheets:

        for coluna in aba.columns:

            maior = 0
            letra = coluna[0].column_letter

            for celula in coluna:

                try:
                    tamanho = len(str(celula.value))

                    if tamanho > maior:
                        maior = tamanho

                except:
                    pass

            aba.column_dimensions[
                letra
            ].width = maior + 3

    response = HttpResponse(
        content_type=(
            'application/vnd.openxmlformats-officedocument.'
            'spreadsheetml.sheet'
        )
    )

    response['Content-Disposition'] = (
        'attachment; filename=relatorio_executivo_ctos.xlsx'
    )

    workbook.save(response)

    return response


@login_required
def exportar_ctos_pdf(request):

    response = HttpResponse(
        content_type='application/pdf'
    )

    response['Content-Disposition'] = (
        'attachment; filename=relatorio_executivo_ctos.pdf'
    )

    documento = SimpleDocTemplate(
        response,
        pagesize=letter
    )

    elementos = []

    estilos = getSampleStyleSheet()

    elementos.append(
        Paragraph(
            "<b>RELATÓRIO EXECUTIVO DE CTOs</b>",
            estilos['Title']
        )
    )

    elementos.append(
        Paragraph(
            f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
            estilos['Normal']
        )
    )

    elementos.append(Spacer(1, 20))

    ctos = CTO.objects.all()

    total_ctos = ctos.count()

    total_portas = sum(
        cto.portas_total for cto in ctos
    )

    total_ocupadas = sum(
        cto.portas_ocupadas for cto in ctos
    )

    total_livres = (
        total_portas - total_ocupadas
    )

    if total_portas > 0:
        taxa_ocupacao = round(
            (total_ocupadas / total_portas) * 100,
            1
        )
    else:
        taxa_ocupacao = 0

    elementos.append(
        Paragraph(
            "<b>RESUMO EXECUTIVO</b>",
            estilos['Heading2']
        )
    )

    resumo = [
        f"Total de CTOs: {total_ctos}",
        f"Total de Portas: {total_portas}",
        f"Portas Ocupadas: {total_ocupadas}",
        f"Portas Livres: {total_livres}",
        f"Taxa Geral de Ocupação: {taxa_ocupacao}%"
    ]

    for item in resumo:

        elementos.append(
            Paragraph(
                item,
                estilos['Normal']
            )
        )

    elementos.append(Spacer(1, 20))

    elementos.append(
        Paragraph(
            "<b>RANKING DE OCUPAÇÃO</b>",
            estilos['Heading2']
        )
    )

    ranking = []

    for cto in ctos:

        if cto.portas_total > 0:

            percentual = round(
                (
                    cto.portas_ocupadas /
                    cto.portas_total
                ) * 100,
                1
            )

        else:

            percentual = 0

        ranking.append([
            cto.nome,
            str(cto.portas_ocupadas),
            str(cto.portas_total),
            f"{percentual}%"
        ])

    ranking.sort(
        key=lambda linha: float(
            linha[3].replace('%', '')
        ),
        reverse=True
    )

    dados = [[
        'CTO',
        'Ocupadas',
        'Totais',
        'Percentual'
    ]]

    dados.extend(ranking)

    tabela = Table(
        dados,
        colWidths=[220, 90, 90, 90]
    )

    tabela.setStyle(
        TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ])
    )

    elementos.append(tabela)

    documento.build(elementos)

    return response


from django.contrib.auth.decorators import user_passes_test


@login_required
@user_passes_test(
    lambda u: u.groups.filter(
        name='Administrador'
    ).exists()
)
def excluir_cto(
    request,
    cto_id
):

    cto = get_object_or_404(
        CTO,
        id=cto_id
    )

    quantidade_clientes = Cliente.objects.filter(
        cto=cto
    ).count()

    if quantidade_clientes > 0:

        return render(
            request,
            'ctos/erro_exclusao.html',
            {
                'cto': cto,
                'quantidade_clientes': quantidade_clientes,
            }
        )

    if request.method == 'POST':

        HistoricoMovimentacao.objects.create(
            usuario=request.user.username,
            cliente_nome="-",
            cto_nome=cto.nome,
            porta=0,
            acao="CTO EXCLUÍDA",
            observacao=f"Rua: {cto.rua}"
        )

        cto.delete()

        return redirect(
            '/ctos/'
        )

    return render(
        request,
        'ctos/excluir_cto.html',
        {
            'cto': cto
        }
    )