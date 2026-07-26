import json
import shutil
from datetime import date, datetime, timedelta

import openpyxl
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import cm

from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.mixins import LoginRequiredMixin
import re
from django.db import models, connections
from django.http import HttpResponse, JsonResponse, FileResponse, Http404
from .backup_utils import fazer_backup, listar_backups, BACKUPS_DIR
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse_lazy
from django.utils import timezone
from django.views.generic import ListView, CreateView, UpdateView, DetailView

from django.db.models import Sum, Q

from .models import Plano, CTO, Cliente, Chamado, ContaPagar, ChamadoAnexo, LogAtividade, Pagamento, MovimentacaoReceita, DebitoCongelado, Material, MovimentacaoEstoque, JornadaTrabalho, RegistroPonto, AbonoPonto, LiberacaoExtraPonto
from .forms import ClienteForm, ChamadoForm, ContaPagarForm, CTOForm, PlanoForm, UsuarioCreateForm, UsuarioUpdateForm, DebitoCongeladoForm, NegociarDebitoForm, MaterialForm, EntradaEstoqueForm, SaidaEstoqueForm, JornadaForm, PontoLiberarForm, AbonoForm, LiberacaoExtraForm
from .decorators import somente_operacao, somente_admin
from .mixins import SomenteAdminMixin, SomenteOperacaoMixin
from django.contrib.auth.models import Group, Permission
from .utils import normalizar, rotulo_permissao, MODELO_LABELS, ACAO_LABELS, proximo_tipo_ponto, tem_entrada_hoje, resumo_ponto_dia
from accounts.models import User


def _somente_tecnico(user):
    return user.is_authenticated and user.role == "tecnico"


# ---------------------------------------------------------------------------
# DASHBOARD
# ---------------------------------------------------------------------------

@login_required
def dashboard(request):
    if request.user.role == "tecnico":
        return redirect("meus_chamados")

    clientes = Cliente.objects.all()
    ctos = CTO.objects.all()
    total_portas = sum(c.capacidade for c in ctos) or 0
    portas_ocupadas = sum(c.portas_ocupadas() for c in ctos) or 0

    context = {
        "total_ctos": ctos.count(),
        "total_clientes": clientes.count(),
        "ctos_ocupadas": sum(1 for c in ctos if c.portas_ocupadas() > 0),
        "ctos_lotadas": [c for c in ctos if c.esta_lotada()],
        "total_portas": total_portas,
        "portas_ocupadas": portas_ocupadas,
        "portas_livres": total_portas - portas_ocupadas,
        "rede_livre_pct": round(((total_portas - portas_ocupadas) / total_portas) * 100, 1) if total_portas else 0,
        "rede_ocupada_pct": round((portas_ocupadas / total_portas) * 100, 1) if total_portas else 0,
        "clientes_recentes": clientes.select_related("plano", "cto")[:5],
        "chamados_abertos": Chamado.objects.exclude(status__in=["concluido", "cancelado"]).count(),
        "chamados_em_atendimento": Chamado.objects.filter(status="andamento").count(),
        "inadimplentes": clientes.filter(status="inadimplente").count(),
    }

    if request.user.role == "admin":
        context["logs_recentes"] = LogAtividade.objects.select_related("usuario")[:15]

    return render(request, "core/dashboard.html", context)


@login_required
def cliente_busca_json(request):
    """Busca livre de cliente por nome, login PPPoE, rua, telefone ou CPF —
    usada pelo campo de busca ao abrir um chamado."""
    termo = request.GET.get("q", "").strip()
    if len(termo) < 2:
        return JsonResponse({"resultados": []})

    alvo = normalizar(termo)
    candidatos = Cliente.objects.exclude(status="inativo").select_related("cto")[:500]
    encontrados = []
    for c in candidatos:
        campos = [c.nome, c.login_pppoe, c.logradouro, c.telefone, c.cpf_cnpj]
        if any(alvo in normalizar(campo) for campo in campos if campo):
            encontrados.append({
                "id": c.id,
                "nome": c.nome,
                "detalhe": f"{c.telefone or '—'} · {c.endereco_completo()}",
            })
        if len(encontrados) >= 15:
            break
    return JsonResponse({"resultados": encontrados})


# ---------------------------------------------------------------------------
# CLIENTES (admin + operador)
# ---------------------------------------------------------------------------

class ClienteListView(SomenteOperacaoMixin, ListView):
    model = Cliente
    template_name = "core/cliente_list.html"
    context_object_name = "clientes"
    paginate_by = 25

    def get_queryset(self):
        qs = super().get_queryset().select_related("plano", "cto")
        status = self.request.GET.get("status")
        if status:
            qs = qs.filter(status=status)

        busca = self.request.GET.get("q")
        if busca:
            alvo = normalizar(busca)
            ids_encontrados = [
                c.id for c in qs
                if alvo in normalizar(c.nome)
                or alvo in normalizar(c.cpf_cnpj)
                or alvo in normalizar(c.telefone)
                or alvo in normalizar(c.bairro)
                or alvo in normalizar(c.cto.codigo if c.cto else "")
            ]
            qs = qs.filter(id__in=ids_encontrados)
        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["busca"] = self.request.GET.get("q", "")
        context["status_filtro"] = self.request.GET.get("status", "")
        context["status_choices"] = Cliente.STATUS_CHOICES
        return context


class ClienteCreateView(SomenteOperacaoMixin, CreateView):
    model = Cliente
    form_class = ClienteForm
    template_name = "core/cliente_form.html"
    success_url = reverse_lazy("cliente_list")

    def form_valid(self, form):
        response = super().form_valid(form)
        MovimentacaoReceita.objects.create(
            cliente=self.object, tipo="novo_cliente",
            valor_anterior=0, valor_novo=self.object.valor_mensal(),
            criado_por=self.request.user,
        )
        messages.success(self.request, "Cliente cadastrado com sucesso!")
        return response


class ClienteUpdateView(SomenteOperacaoMixin, UpdateView):
    model = Cliente
    form_class = ClienteForm
    template_name = "core/cliente_form.html"
    success_url = reverse_lazy("cliente_list")

    def form_valid(self, form):
        cliente_antigo = Cliente.objects.get(pk=self.object.pk)
        valor_antigo = cliente_antigo.valor_mensal()
        plano_antigo_id = cliente_antigo.plano_id

        response = super().form_valid(form)

        valor_novo = self.object.valor_mensal()
        if self.object.plano_id != plano_antigo_id and valor_novo != valor_antigo:
            tipo = "upgrade" if valor_novo > valor_antigo else "downgrade"
            MovimentacaoReceita.objects.create(
                cliente=self.object, tipo=tipo,
                valor_anterior=valor_antigo, valor_novo=valor_novo,
                criado_por=self.request.user,
            )
        messages.success(self.request, "Dados do cliente atualizados!")
        return response


@user_passes_test(lambda u: u.is_authenticated and u.role in ("admin", "operador"))
def cliente_delete(request, pk):
    cliente = get_object_or_404(Cliente, pk=pk)
    if request.method == "POST":
        nome = cliente.nome
        cliente.delete()
        messages.success(request, f"Cliente \"{nome}\" foi excluído permanentemente do sistema.")
        return redirect("cliente_list")
    return render(request, "core/cliente_confirm_delete.html", {"cliente": cliente})


@login_required
def cto_portas_livres(request, pk):
    cto = get_object_or_404(CTO, pk=pk)
    livres = cto.portas_disponiveis()
    cliente_id = request.GET.get("cliente_id")
    if cliente_id:
        try:
            cliente = Cliente.objects.get(pk=cliente_id, cto_id=pk)
            if cliente.porta and cliente.porta not in livres:
                livres = sorted(livres + [cliente.porta])
        except Cliente.DoesNotExist:
            pass
    return JsonResponse({"portas": livres})


@user_passes_test(lambda u: u.is_authenticated and u.role in ("admin", "operador"))
def cliente_cancelar(request, pk):
    cliente = get_object_or_404(Cliente, pk=pk)
    if request.method == "POST":
        motivo = request.POST.get("motivo_cancelamento", "").strip()
        if not motivo:
            messages.error(request, "É obrigatório informar o motivo do cancelamento.")
            return redirect("cliente_cancelar", pk=pk)

        valor_anterior = cliente.valor_mensal()
        cliente.status = "cancelado"
        cliente.motivo_cancelamento = motivo
        cliente.data_cancelamento = timezone.now().date()
        cliente.save()

        MovimentacaoReceita.objects.create(
            cliente=cliente, tipo="cancelamento",
            valor_anterior=valor_anterior, valor_novo=0,
            criado_por=request.user,
        )
        messages.success(request, f"Cliente \"{cliente.nome}\" marcado como cancelado.")
        return redirect("cliente_list")
    return render(request, "core/cliente_cancelar.html", {"cliente": cliente})


class ClienteCanceladosListView(SomenteOperacaoMixin, ListView):
    model = Cliente
    template_name = "core/cliente_cancelados.html"
    context_object_name = "clientes"

    def get_queryset(self):
        return Cliente.objects.filter(status="cancelado").select_related("plano", "cto").order_by("-data_cancelamento")


@somente_operacao
def cliente_historico(request, pk):
    cliente = get_object_or_404(Cliente, pk=pk)
    chamados = (
        Chamado.objects.filter(cliente=cliente)
        .select_related("tecnico", "aberto_por")
        .order_by("-criado_em")
    )
    pagamentos = cliente.pagamentos.order_by("-mes_referencia")

    chamados_por_ano = {}
    for c in chamados:
        ano = c.criado_em.year
        chamados_por_ano[ano] = chamados_por_ano.get(ano, 0) + 1

    context = {
        "cliente": cliente,
        "chamados": chamados,
        "pagamentos": pagamentos,
        "total_chamados": chamados.count(),
        "chamados_por_ano": sorted(chamados_por_ano.items(), reverse=True),
    }
    return render(request, "core/cliente_historico.html", context)


# ---------------------------------------------------------------------------
# CTOs (leitura: qualquer perfil logado / escrita: admin+operador)
# ---------------------------------------------------------------------------

class CTOListView(LoginRequiredMixin, ListView):
    model = CTO
    template_name = "core/cto_list.html"
    context_object_name = "ctos"

    def get_queryset(self):
        qs = super().get_queryset()
        busca = self.request.GET.get("q")
        if busca:
            alvo = normalizar(busca)
            ids_encontrados = [
                c.id for c in qs
                if alvo in normalizar(c.codigo) or alvo in normalizar(c.bairro) or alvo in normalizar(c.endereco)
                or any(alvo in normalizar(r) for r in c.lista_ruas())
            ]
            qs = qs.filter(id__in=ids_encontrados)
        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["busca"] = self.request.GET.get("q", "")
        return context


class CTODetailView(LoginRequiredMixin, DetailView):
    model = CTO
    context_object_name = "cto"

    def get_template_names(self):
        if self.request.user.role == "tecnico":
            return ["core/cto_detail_tecnico.html"]
        return ["core/cto_detail.html"]

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["clientes_conectados"] = self.object.clientes.exclude(status="inativo")
        return context


class CTOCreateView(SomenteOperacaoMixin, CreateView):
    model = CTO
    form_class = CTOForm
    template_name = "core/cto_form.html"
    success_url = reverse_lazy("cto_list")

    def form_valid(self, form):
        messages.success(self.request, "CTO cadastrada com sucesso!")
        return super().form_valid(form)


class CTOUpdateView(SomenteOperacaoMixin, UpdateView):
    model = CTO
    form_class = CTOForm
    template_name = "core/cto_form.html"
    success_url = reverse_lazy("cto_list")

    def form_valid(self, form):
        messages.success(self.request, "CTO atualizada com sucesso!")
        return super().form_valid(form)


# ---------------------------------------------------------------------------
# PLANOS (admin + operador)
# ---------------------------------------------------------------------------

class PlanoListView(SomenteOperacaoMixin, ListView):
    model = Plano
    template_name = "core/plano_list.html"
    context_object_name = "planos"


class PlanoCreateView(SomenteOperacaoMixin, CreateView):
    model = Plano
    form_class = PlanoForm
    template_name = "core/plano_form.html"
    success_url = reverse_lazy("plano_list")

    def form_valid(self, form):
        messages.success(self.request, "Plano cadastrado com sucesso!")
        return super().form_valid(form)


# ---------------------------------------------------------------------------
# CHAMADOS — visão admin/operador (gestão completa)
# ---------------------------------------------------------------------------

class ChamadoListView(SomenteOperacaoMixin, ListView):
    model = Chamado
    template_name = "core/chamado_list.html"
    context_object_name = "chamados"

    def get_queryset(self):
        qs = super().get_queryset().exclude(status__in=["concluido", "cancelado"]).select_related("cliente", "tecnico")
        status = self.request.GET.get("status")
        tecnico_id = self.request.GET.get("tecnico")
        if status:
            qs = qs.filter(status=status)
        if tecnico_id:
            qs = qs.filter(tecnico_id=tecnico_id)
        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["tecnicos"] = User.objects.filter(role="tecnico")
        context["status_choices"] = [c for c in Chamado.STATUS_CHOICES if c[0] not in ("concluido", "cancelado")]
        context["prioridade_choices"] = Chamado.PRIORIDADE_CHOICES
        return context


def _info_retorno_cliente(cliente, excluir_pk=None):
    """Verifica se o cliente teve chamado nos últimos 30 dias e devolve os detalhes,
    pra avisar o operador antes mesmo de abrir o chamado novo."""
    limite = timezone.now() - timedelta(days=30)
    qs = Chamado.objects.filter(cliente=cliente, criado_em__gte=limite).exclude(status="cancelado")
    if excluir_pk:
        qs = qs.exclude(pk=excluir_pk)
    anterior = qs.order_by("-criado_em").first()
    if not anterior:
        return None
    return {
        "tipo": anterior.get_tipo_display(),
        "descricao": anterior.descricao or "",
        "data": anterior.criado_em.strftime("%d/%m/%Y às %H:%M"),
        "tecnico": anterior.tecnico.get_full_name() if anterior.tecnico else None,
        "status": anterior.get_status_display(),
    }


@login_required
def cliente_verificar_retorno(request, pk):
    cliente = get_object_or_404(Cliente, pk=pk)
    info = _info_retorno_cliente(cliente)
    return JsonResponse({"retorno": info is not None, "info": info})


class ChamadoCreateView(SomenteOperacaoMixin, CreateView):
    model = Chamado
    form_class = ChamadoForm
    template_name = "core/chamado_form.html"
    success_url = reverse_lazy("chamado_list")

    def get_initial(self):
        initial = super().get_initial()
        cliente_id = self.request.GET.get("cliente")
        if cliente_id:
            initial["cliente"] = cliente_id
        return initial

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        cliente_id = self.request.GET.get("cliente")
        if cliente_id:
            context["cliente_pre_selecionado"] = Cliente.objects.filter(pk=cliente_id).first()
        return context

    def form_valid(self, form):
        form.instance.aberto_por = self.request.user
        response = super().form_valid(form)
        chamado = self.object

        limite = timezone.now() - timedelta(days=30)
        anterior = (
            Chamado.objects.filter(cliente=chamado.cliente, criado_em__gte=limite)
            .exclude(pk=chamado.pk)
            .order_by("-criado_em")
            .first()
        )
        if anterior:
            chamado.eh_retorno = True
            chamado.tecnico_ultimo_atendimento = anterior.tecnico
            chamado.prioridade = "extrema"
            chamado.save()
            messages.warning(
                self.request,
                f"Atenção: {chamado.cliente.nome} já abriu um chamado nos últimos 30 dias "
                f"(atendido por {anterior.tecnico or 'ninguém ainda'}). "
                f"Marcado automaticamente como RETORNO com prioridade Extrema.",
            )
        return response


@user_passes_test(lambda u: u.is_authenticated and u.role in ("admin", "operador"))
def reatribuir_chamado(request, pk):
    chamado = get_object_or_404(Chamado, pk=pk)
    if request.method == "POST":
        tecnico_id = request.POST.get("tecnico")
        chamado.tecnico_id = tecnico_id or None
        chamado.save()
    return redirect("chamado_list")


RANK_PRIORIDADE = {"extrema": 4, "alta": 3, "media": 2, "baixa": 1}


@user_passes_test(lambda u: u.is_authenticated and u.role in ("admin", "operador"))
def alterar_prioridade_chamado(request, pk):
    chamado = get_object_or_404(Chamado, pk=pk)
    if request.method == "POST":
        nova = request.POST.get("prioridade", chamado.prioridade)
        justificativa = request.POST.get("justificativa", "").strip()
        antiga = chamado.prioridade

        if nova != antiga:
            reducao = RANK_PRIORIDADE.get(nova, 0) < RANK_PRIORIDADE.get(antiga, 0)
            if reducao and not justificativa:
                messages.error(request, "É necessário justificar quando a prioridade é reduzida.")
                return redirect("chamado_list")

            chamado.prioridade = nova
            chamado.save()

            detalhes = f"Chamado #{chamado.id} ({chamado.cliente.nome}): {antiga} → {nova}"
            if justificativa:
                detalhes += f" — Justificativa: {justificativa}"
            LogAtividade.objects.create(usuario=request.user, acao="Prioridade do chamado alterada", detalhes=detalhes)
    return redirect("chamado_list")


@user_passes_test(lambda u: u.is_authenticated and u.role in ("admin", "operador"))
def cancelar_chamado(request, pk):
    chamado = get_object_or_404(Chamado, pk=pk)
    chamado.status = "cancelado"
    chamado.save()
    messages.success(request, f"Chamado #{chamado.id} cancelado.")
    return redirect("chamado_list")


@user_passes_test(lambda u: u.is_authenticated and u.role in ("admin", "operador"))
def voltar_chamado_aberto(request, pk):
    """Usado quando algo aconteceu com o técnico e o chamado precisa ser reaberto
    pra outro pegar."""
    chamado = get_object_or_404(Chamado, pk=pk)
    chamado.status = "aberto"
    chamado.tecnico = None
    chamado.pego_em = None
    chamado.save()
    messages.success(request, f"Chamado #{chamado.id} voltou para Aberto e está disponível de novo.")
    return redirect("chamado_list")


@user_passes_test(lambda u: u.is_authenticated and u.role in ("admin", "operador"))
def finalizar_chamado_operador(request, pk):
    """O operador finaliza um chamado no lugar do técnico (ex: técnico com problema)."""
    chamado = get_object_or_404(Chamado, pk=pk)
    if request.method == "POST":
        observacao = request.POST.get("observacao_fechamento", "").strip()
        if not observacao:
            messages.error(request, "Descreva o motivo antes de finalizar o chamado.")
            return redirect("finalizar_chamado_operador", pk=pk)
        chamado.observacao_fechamento = observacao
        chamado.status = "concluido"
        chamado.concluido_em = timezone.now()
        chamado.finalizado_por = request.user
        if not chamado.pego_em:
            chamado.pego_em = chamado.concluido_em
        chamado.save()
        if chamado.tecnico and chamado.tecnico != request.user:
            messages.success(
                request,
                f"Chamado #{chamado.id} finalizado por você — era o chamado do técnico "
                f"{chamado.tecnico.get_full_name() or chamado.tecnico.username}.",
            )
        else:
            messages.success(request, f"Chamado #{chamado.id} finalizado por você.")
        return redirect("chamado_list")
    return render(request, "core/chamado_finalizar_operador.html", {"chamado": chamado})


class ChamadoFinalizadosListView(SomenteOperacaoMixin, ListView):
    model = Chamado
    template_name = "core/chamado_finalizados.html"
    context_object_name = "chamados"

    def get_queryset(self):
        return Chamado.objects.filter(status="concluido").select_related("cliente", "tecnico")


# ---------------------------------------------------------------------------
# ÁREA DO TÉCNICO
# ---------------------------------------------------------------------------

@login_required
def meus_chamados(request):
    chamados = (
        Chamado.objects.filter(tecnico=request.user)
        .exclude(status__in=["cancelado", "concluido"])
        .select_related("cliente")
    )
    context = {
        "chamados": chamados,
        "abertos": chamados.filter(status="aberto").count(),
        "andamento": chamados.filter(status="andamento").count(),
        "concluidos": Chamado.objects.filter(tecnico=request.user, status="concluido").count(),
    }
    return render(request, "core/meus_chamados.html", context)


@user_passes_test(_somente_tecnico)
def chamados_disponiveis(request):
    """Lista todos os chamados em aberto. Os já assumidos por um técnico aparecem
    destacados/bloqueados, para nenhum outro técnico pegar o mesmo chamado."""
    if not tem_entrada_hoje(request.user):
        messages.warning(request, "Você precisa bater o ponto de entrada antes de ver os chamados disponíveis.")
        return redirect("ponto_bater")
    chamados = (
        Chamado.objects.filter(status__in=["aberto", "andamento"])
        .exclude(status="cancelado")
        .select_related("cliente", "tecnico")
    )
    return render(request, "core/chamados_disponiveis.html", {"chamados": chamados})


@user_passes_test(_somente_tecnico)
def pegar_chamado(request, pk):
    if not tem_entrada_hoje(request.user):
        messages.warning(request, "Você precisa bater o ponto de entrada antes de pegar um chamado.")
        return redirect("ponto_bater")
    chamado = get_object_or_404(Chamado, pk=pk)

    ja_tem_ativo = Chamado.objects.filter(
        tecnico=request.user, status__in=["aberto", "andamento"]
    ).exclude(pk=pk).exists()
    if ja_tem_ativo:
        messages.warning(request, "Você já tem um chamado em atendimento. Finalize-o antes de pegar outro.")
        return redirect("chamados_disponiveis")

    if chamado.tecnico_id is None:
        chamado.tecnico = request.user
        chamado.pego_em = timezone.now()
        chamado.save()
        messages.success(request, f"Chamado #{chamado.id} atribuído a você. Bata uma foto pra iniciar o atendimento.")
        return redirect("meus_chamados")
    else:
        messages.warning(request, "Esse chamado já foi assumido por outro técnico.")
    return redirect("chamados_disponiveis")


@login_required
def avancar_chamado(request, pk):
    """Passo 'Iniciar atendimento' (aguardando início -> em andamento). Exige uma
    foto tirada na hora (câmera do celular) como comprovante de chegada no cliente,
    e marca o horário exato de início pra calcular quanto tempo o técnico ficou lá."""
    chamado = get_object_or_404(Chamado, pk=pk, tecnico=request.user)
    if chamado.status != "aberto":
        return redirect("meus_chamados")

    if request.method == "POST":
        foto = request.FILES.get("foto_inicio")
        if not foto:
            messages.error(request, "Tire uma foto pra iniciar o atendimento.")
            return redirect("avancar_chamado", pk=pk)
        chamado.foto_inicio = foto
        chamado.atendimento_iniciado_em = timezone.now()
        chamado.status = "andamento"
        chamado.save()
        messages.success(request, f"Atendimento iniciado às {timezone.localtime().strftime('%H:%M')}.")
        return redirect("meus_chamados")

    return render(request, "core/chamado_iniciar.html", {"chamado": chamado})


@login_required
def finalizar_chamado(request, pk):
    chamado = get_object_or_404(Chamado, pk=pk, tecnico=request.user)
    if request.method == "POST":
        observacao = request.POST.get("observacao_fechamento", "").strip()
        if not observacao:
            messages.error(request, "Descreva o que foi feito antes de concluir o chamado.")
            return redirect("finalizar_chamado", pk=pk)

        chamado.observacao_fechamento = observacao
        chamado.status = "concluido"
        chamado.concluido_em = timezone.now()
        chamado.finalizado_por = request.user
        chamado.save()

        for foto in request.FILES.getlist("fotos"):
            ChamadoAnexo.objects.create(chamado=chamado, imagem=foto)

        messages.success(request, f"Chamado #{chamado.id} concluído com sucesso!")
        return redirect("meus_chamados")

    return render(request, "core/chamado_finalizar.html", {"chamado": chamado})


def garantir_contas_recorrentes(ano_alvo, mes_alvo):
    """Garante que as contas fixas e parceladas já tenham uma cópia criada até o mês
    que está sendo visualizado na tela (não só até a data real de hoje)."""
    alvo = date(ano_alvo, mes_alvo, 1)
    raizes = ContaPagar.objects.filter(gerada_de__isnull=True).filter(Q(recorrente=True) | Q(total_parcelas__gt=1))

    for raiz in raizes:
        serie = list(ContaPagar.objects.filter(Q(pk=raiz.pk) | Q(gerada_de=raiz)).order_by("vencimento"))
        ultima = serie[-1]
        quantidade_existente = len(serie)
        seguranca = 0

        while ultima.vencimento.replace(day=1) < alvo and seguranca < 36:
            if not raiz.recorrente and quantidade_existente >= raiz.total_parcelas:
                break  # já geramos todas as parcelas dessa série, não cria mais

            mes_p = ultima.vencimento.month % 12 + 1
            ano_p = ultima.vencimento.year + (1 if ultima.vencimento.month == 12 else 0)
            dia_p = min(ultima.vencimento.day, 28)
            nova_data = date(ano_p, mes_p, dia_p)

            ultima = ContaPagar.objects.create(
                descricao=raiz.descricao, valor=ultima.valor, vencimento=nova_data,
                status="pendente", recorrente=raiz.recorrente, forma_pagamento=raiz.forma_pagamento,
                parcela_atual=(quantidade_existente + 1) if not raiz.recorrente else 1,
                total_parcelas=raiz.total_parcelas,
                gerada_de=raiz,
            )
            quantidade_existente += 1
            seguranca += 1


def _redirect_contas_pagar(request):
    mes = request.POST.get("mes") or request.GET.get("mes")
    ano = request.POST.get("ano") or request.GET.get("ano")
    url = "/financeiro/contas-pagar/"
    if mes and ano:
        url += f"?mes={mes}&ano={ano}"
    return redirect(url)


# ---------------------------------------------------------------------------
# FINANCEIRO (admin + operador)
# ---------------------------------------------------------------------------

def _serie_completa_paga(conta):
    """Verifica se TODAS as parcelas dessa compra já foram pagas (ou se é uma
    conta simples de uma vez só, já paga). Contas fixas (recorrentes) nunca
    'terminam', então nunca entram aqui."""
    if conta.recorrente:
        return False
    raiz = conta.gerada_de or conta
    serie = ContaPagar.objects.filter(Q(pk=raiz.pk) | Q(gerada_de=raiz))
    if serie.count() < raiz.total_parcelas:
        return False
    return not serie.exclude(status="pago").exists()


@somente_admin
def contas_pagas_view(request):
    resultados = []
    raizes = ContaPagar.objects.filter(gerada_de__isnull=True, recorrente=False)
    for raiz in raizes:
        serie = ContaPagar.objects.filter(Q(pk=raiz.pk) | Q(gerada_de=raiz))
        if serie.count() >= raiz.total_parcelas and not serie.exclude(status="pago").exists():
            ultima = serie.order_by("-vencimento").first()
            total_pago = sum((c.valor for c in serie), 0)
            resultados.append({
                "descricao": raiz.descricao,
                "total_parcelas": raiz.total_parcelas,
                "total_pago": total_pago,
                "ultimo_vencimento": ultima.vencimento,
                "nota_fiscal": ultima.nota_fiscal or raiz.nota_fiscal,
                "conta_pk": ultima.pk,
            })
    resultados.sort(key=lambda r: r["ultimo_vencimento"], reverse=True)
    return render(request, "core/contas_pagas.html", {"resultados": resultados})


@somente_admin
def contas_pagar_view(request):
    hoje = timezone.now().date()
    ano = int(request.GET.get("ano", hoje.year))
    mes = int(request.GET.get("mes", hoje.month))

    garantir_contas_recorrentes(ano, mes)

    contas_pagar_bruto = ContaPagar.objects.filter(vencimento__year=ano, vencimento__month=mes)
    contas_pagar = [c for c in contas_pagar_bruto if not _serie_completa_paga(c)]
    total_pagar = sum((c.valor for c in contas_pagar if c.status != "pago"), 0)
    total_pago = sum((c.valor for c in contas_pagar if c.status == "pago"), 0)

    context = {
        "contas_pagar": contas_pagar,
        "total_pagar": total_pagar,
        "total_pago": total_pago,
        "mes": mes,
        "ano": ano,
        "mes_ref": date(ano, mes, 1),
        "meses_opcoes": [(i, MESES_PT[i]) for i in range(1, 13)],
        "anos_opcoes": list(range(hoje.year - 1, hoje.year + 2)),
    }
    return render(request, "core/contas_pagar.html", context)


@somente_admin
def debitos_congelados_list(request):
    debitos = DebitoCongelado.objects.filter(negociado=False)
    negociados = DebitoCongelado.objects.filter(negociado=True).select_related("conta_pagar_gerada")
    total_congelado = sum((d.valor for d in debitos), 0)
    return render(request, "core/debitos_congelados.html", {
        "debitos": debitos, "negociados": negociados, "total_congelado": total_congelado,
    })


@somente_admin
def debito_congelado_create(request):
    if request.method == "POST":
        form = DebitoCongeladoForm(request.POST)
        if form.is_valid():
            debito = form.save(commit=False)
            debito.criado_por = request.user
            debito.save()
            messages.success(request, "Débito congelado cadastrado — não entra no somatório do mês até ser negociado.")
            return redirect("debitos_congelados")
    else:
        form = DebitoCongeladoForm()
    return render(request, "core/debito_congelado_form.html", {"form": form})


@somente_admin
def debito_congelado_negociar(request, pk):
    debito = get_object_or_404(DebitoCongelado, pk=pk, negociado=False)
    if request.method == "POST":
        form = NegociarDebitoForm(request.POST)
        if form.is_valid():
            valor_parcela = form.cleaned_data["valor_parcela"]
            parcelas = form.cleaned_data["parcelas"]
            primeiro_venc = form.cleaned_data["primeiro_vencimento"]
            forma = form.cleaned_data["forma_pagamento"]

            conta = ContaPagar.objects.create(
                descricao=f"{debito.descricao} (negociado)", valor=valor_parcela, vencimento=primeiro_venc,
                status="pendente", recorrente=False, forma_pagamento=forma,
                parcela_atual=1, total_parcelas=parcelas if forma != "avista" else 1,
            )
            debito.negociado = True
            debito.negociado_em = timezone.now()
            debito.conta_pagar_gerada = conta
            debito.save()
            messages.success(
                request,
                f"Débito \"{debito.descricao}\" negociado — já entrou em Contas a Pagar, começando em "
                f"{primeiro_venc.strftime('%d/%m/%Y')}.",
            )
            return redirect("debitos_congelados")
    else:
        form = NegociarDebitoForm(initial={"valor_parcela": debito.valor, "parcelas": 1})
    return render(request, "core/debito_congelado_negociar.html", {"form": form, "debito": debito})


@somente_admin
def debito_congelado_delete(request, pk):
    debito = get_object_or_404(DebitoCongelado, pk=pk)
    if request.method == "POST":
        descricao = debito.descricao
        debito.delete()
        messages.success(request, f"Débito \"{descricao}\" excluído.")
        return redirect("debitos_congelados")
    return render(request, "core/debito_congelado_confirm_delete.html", {"debito": debito})


def _dias_em_atraso(cliente):
    """Calcula há quantos dias o cliente está sem pagar, olhando o mês seguinte
    ao último pagamento registrado (ou desde a instalação, se nunca pagou)."""
    hoje = timezone.now().date()
    ultimo_pagamento = cliente.pagamentos.order_by("-mes_referencia").first()

    if ultimo_pagamento:
        prox_mes = ultimo_pagamento.mes_referencia.month % 12 + 1
        prox_ano = ultimo_pagamento.mes_referencia.year + (1 if ultimo_pagamento.mes_referencia.month == 12 else 0)
    else:
        base = cliente.data_ativacao or cliente.criado_em.date()
        prox_mes, prox_ano = base.month, base.year

    dia = min(cliente.dia_vencimento or 10, 28)
    data_vencimento = date(prox_ano, prox_mes, dia)
    return max((hoje - data_vencimento).days, 0)


def _formatar_atraso(dias):
    if dias <= 0:
        return "vence em breve"
    meses, resto = divmod(dias, 30)
    if meses > 0 and resto > 0:
        return f"{meses} {'mês' if meses == 1 else 'meses'} e {resto} {'dia' if resto == 1 else 'dias'} em aberto"
    if meses > 0:
        return f"{meses} {'mês' if meses == 1 else 'meses'} em aberto"
    return f"{dias} {'dia' if dias == 1 else 'dias'} em aberto"


@somente_operacao
def clientes_inadimplentes_view(request):
    clientes = Cliente.objects.filter(status="inadimplente").select_related("plano")
    dados = []
    for c in clientes:
        dias = _dias_em_atraso(c)
        dados.append({"cliente": c, "dias": dias, "atraso_texto": _formatar_atraso(dias)})
    dados.sort(key=lambda d: d["dias"], reverse=True)
    return render(request, "core/clientes_inadimplentes.html", {"dados": dados})


@somente_admin
def financeiro(request):
    hoje = timezone.now().date()
    ano = int(request.GET.get("ano", hoje.year))
    mes = int(request.GET.get("mes", hoje.month))
    mes_ref = date(ano, mes, 1)

    garantir_contas_recorrentes(ano, mes)

    clientes = Cliente.objects.exclude(status="inativo").select_related("plano")
    esperado_mes = sum((c.valor_mensal() for c in clientes), 0)

    pagamentos_mes = Pagamento.objects.filter(mes_referencia=mes_ref)
    ids_pagos = set(pagamentos_mes.values_list("cliente_id", flat=True))
    recebido_mes = pagamentos_mes.aggregate(total=Sum("valor"))["total"] or 0
    falta_receber_mes = esperado_mes - recebido_mes

    clientes_status = []
    for c in clientes:
        eh_instalacao = (
            c.data_ativacao and c.data_ativacao.year == ano and c.data_ativacao.month == mes
        )
        clientes_status.append({"cliente": c, "pago": c.id in ids_pagos, "eh_instalacao": eh_instalacao and c.id not in ids_pagos})

    total_clientes_mes = clientes.count()
    pagos_count = len(ids_pagos)
    faltam_count = total_clientes_mes - pagos_count

    inadimplentes = clientes.filter(status="inadimplente")
    total_inadimplencia = sum((c.valor_mensal() for c in inadimplentes), 0)
    contas_pagar = ContaPagar.objects.filter(vencimento__year=ano, vencimento__month=mes)
    total_pagar = sum((c.valor for c in contas_pagar.exclude(status="pago")), 0)

    # Movimentação de receita do mês selecionado: quem entrou, quem cancelou, quem trocou de plano
    movs_mes = MovimentacaoReceita.objects.filter(criado_em__year=ano, criado_em__month=mes).select_related("cliente")
    novos = movs_mes.filter(tipo="novo_cliente")
    cancelamentos = movs_mes.filter(tipo="cancelamento")
    upgrades = movs_mes.filter(tipo="upgrade")
    downgrades = movs_mes.filter(tipo="downgrade")

    receita_nova = sum((m.valor_novo for m in novos), 0)
    receita_perdida = sum((m.valor_anterior for m in cancelamentos), 0)
    ajuste_planos = sum((m.diferenca() for m in upgrades), 0) + sum((m.diferenca() for m in downgrades), 0)
    saldo_mes = receita_nova - receita_perdida + ajuste_planos
    percentual_saldo = (saldo_mes / esperado_mes * 100) if esperado_mes else 0
    percentual_perdido = (receita_perdida / esperado_mes * 100) if esperado_mes else 0
    percentual_novo = (receita_nova / esperado_mes * 100) if esperado_mes else 0

    # Fluxo de caixa dos últimos 6 meses: entradas (pagamentos) x saídas (contas)
    base_mes = hoje.replace(day=1)
    meses_labels, entradas, saidas = [], [], []
    for i in range(5, -1, -1):
        ano_i = base_mes.year + ((base_mes.month - i - 1) // 12)
        mes_i = ((base_mes.month - i - 1) % 12) + 1
        ref = base_mes.replace(year=ano_i, month=mes_i)
        entrada_mes = Pagamento.objects.filter(mes_referencia=ref).aggregate(total=Sum("valor"))["total"] or 0
        saida_mes = ContaPagar.objects.filter(vencimento__year=ref.year, vencimento__month=ref.month).aggregate(
            total=Sum("valor")
        )["total"] or 0
        meses_labels.append(MESES_PT[mes_i][:3])
        entradas.append(float(entrada_mes))
        saidas.append(float(saida_mes))

    context = {
        "clientes_status": clientes_status,
        "mes_ref": mes_ref,
        "mes": mes,
        "ano": ano,
        "meses_opcoes": [(i, MESES_PT[i]) for i in range(1, 13)],
        "anos_opcoes": list(range(hoje.year - 1, hoje.year + 2)),
        "esperado_mes": esperado_mes,
        "recebido_mes": recebido_mes,
        "falta_receber_mes": falta_receber_mes,
        "total_clientes_mes": total_clientes_mes,
        "pagos_count": pagos_count,
        "faltam_count": faltam_count,
        "inadimplentes": inadimplentes,
        "total_inadimplencia": total_inadimplencia,
        "contas_pagar": contas_pagar,
        "total_pagar": total_pagar,
        "meses_labels": json.dumps(meses_labels),
        "entradas": json.dumps(entradas),
        "saidas": json.dumps(saidas),
        "novos_count": novos.count(),
        "cancelamentos_count": cancelamentos.count(),
        "receita_nova": receita_nova,
        "receita_perdida": receita_perdida,
        "ajuste_planos": ajuste_planos,
        "saldo_mes": saldo_mes,
        "percentual_saldo": percentual_saldo,
        "percentual_perdido": percentual_perdido,
        "percentual_novo": percentual_novo,
    }
    return render(request, "core/financeiro.html", context)


@somente_admin
def registrar_pagamento(request, pk):
    cliente = get_object_or_404(Cliente, pk=pk)
    hoje = timezone.now().date()
    ano = int(request.POST.get("ano", hoje.year))
    mes = int(request.POST.get("mes", hoje.month))
    mes_ref = date(ano, mes, 1)

    Pagamento.objects.update_or_create(
        cliente=cliente, mes_referencia=mes_ref,
        defaults={"valor": cliente.valor_mensal(), "data_pagamento": hoje, "registrado_por": request.user},
    )
    if cliente.status == "inadimplente":
        cliente.status = "ativo"
        cliente.save()
    messages.success(request, f"Pagamento de {cliente.nome} registrado para {MESES_PT[mes]}/{ano}.")
    return redirect(f"/financeiro/?mes={mes}&ano={ano}")


@somente_admin
def conta_pagar_create(request):
    if request.method == "POST":
        form = ContaPagarForm(request.POST, request.FILES)
        if form.is_valid():
            conta = form.save(commit=False)
            parcelas = form.cleaned_data.get("parcelas") or 1

            if not conta.recorrente and conta.forma_pagamento in ("boleto", "cartao") and parcelas > 1:
                conta.total_parcelas = parcelas
                conta.parcela_atual = 1
            else:
                conta.total_parcelas = 1
                conta.parcela_atual = 1
            conta.save()

            messages.success(request, "Conta a pagar cadastrada com sucesso!")
            return _redirect_contas_pagar(request)
    else:
        form = ContaPagarForm()
    return render(request, "core/conta_pagar_form.html", {
        "form": form, "mes": request.GET.get("mes"), "ano": request.GET.get("ano"),
    })


@somente_admin
def conta_pagar_update(request, pk):
    conta = get_object_or_404(ContaPagar, pk=pk)
    if request.method == "POST":
        form = ContaPagarForm(request.POST, request.FILES, instance=conta)
        if form.is_valid():
            form.save()
            messages.success(request, "Conta a pagar atualizada!")
            return _redirect_contas_pagar(request)
    else:
        form = ContaPagarForm(instance=conta)
    return render(request, "core/conta_pagar_form.html", {
        "form": form, "object": conta, "mes": request.GET.get("mes"), "ano": request.GET.get("ano"),
    })


@somente_admin
def conta_pagar_delete(request, pk):
    conta = get_object_or_404(ContaPagar, pk=pk)
    if request.method == "POST":
        descricao = conta.descricao
        if conta.recorrente or conta.total_parcelas > 1:
            raiz = conta.gerada_de or conta
            serie = ContaPagar.objects.filter(Q(pk=raiz.pk) | Q(gerada_de=raiz))
            qtd = serie.count()
            serie.delete()
            messages.success(request, f"\"{descricao}\" excluída — {qtd} ocorrência(s) removida(s) (a série inteira, não só esse mês).")
        else:
            conta.delete()
            messages.success(request, f"Conta \"{descricao}\" excluída.")
        return _redirect_contas_pagar(request)
    return render(request, "core/conta_pagar_confirm_delete.html", {
        "conta": conta, "mes": request.GET.get("mes"), "ano": request.GET.get("ano"),
    })


@somente_admin
def conta_pagar_marcar_pago(request, pk):
    conta = get_object_or_404(ContaPagar, pk=pk)
    conta.status = "pago"
    conta.save()
    messages.success(request, f"Conta \"{conta.descricao}\" marcada como paga.")
    return _redirect_contas_pagar(request)


@somente_admin
def conta_pagar_desmarcar_pago(request, pk):
    conta = get_object_or_404(ContaPagar, pk=pk)
    conta.status = "pendente"
    conta.save()
    messages.success(request, f"Pagamento de \"{conta.descricao}\" desfeito — voltou para Pendente.")
    return _redirect_contas_pagar(request)


@somente_admin
def financeiro_export_excel(request):
    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "A Receber"
    ws1.append(["Cliente", "Plano", "Valor", "Pago este mês?"])
    for c in Cliente.objects.exclude(status="inativo").select_related("plano"):
        ws1.append([c.nome, str(c.plano) if c.plano else "", float(c.valor_mensal()), "Sim" if c.pago_mes_atual() else "Não"])

    ws2 = wb.create_sheet("Contas a Pagar")
    ws2.append(["Descrição", "Vencimento", "Valor", "Status"])
    for c in ContaPagar.objects.all():
        ws2.append([c.descricao, c.vencimento.strftime("%d/%m/%Y"), float(c.valor), c.get_status_display()])

    ws3 = wb.create_sheet("Inadimplentes")
    ws3.append(["Cliente", "Telefone", "Valor em atraso"])
    for c in Cliente.objects.filter(status="inadimplente"):
        ws3.append([c.nome, c.telefone, float(c.valor_mensal())])

    for ws in (ws1, ws2, ws3):
        for i in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(i)].width = 22

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="financeiro.xlsx"'
    wb.save(response)
    return response


@somente_admin
def financeiro_export_pdf(request):
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="financeiro.pdf"'
    doc = SimpleDocTemplate(response, pagesize=landscape(A4), topMargin=1.5 * cm, bottomMargin=1.5 * cm)
    estilos = getSampleStyleSheet()
    elementos = [Paragraph("Relatório Financeiro - CTO Manager Pro", estilos["Title"]), Spacer(1, 12)]

    elementos.append(Paragraph("A Receber", estilos["Heading2"]))
    dados1 = [["Cliente", "Plano", "Valor", "Pago este mês?"]]
    for c in Cliente.objects.exclude(status="inativo").select_related("plano"):
        dados1.append([c.nome, str(c.plano) if c.plano else "-", f"R$ {c.valor_mensal():.2f}", "Sim" if c.pago_mes_atual() else "Não"])
    t1 = Table(dados1, repeatRows=1)
    t1.setStyle(_estilo_tabela_pdf())
    elementos += [t1, Spacer(1, 16)]

    elementos.append(Paragraph("Contas a Pagar", estilos["Heading2"]))
    dados2 = [["Descrição", "Vencimento", "Valor", "Status"]]
    for c in ContaPagar.objects.all():
        dados2.append([c.descricao, c.vencimento.strftime("%d/%m/%Y"), f"R$ {c.valor:.2f}", c.get_status_display()])
    t2 = Table(dados2, repeatRows=1)
    t2.setStyle(_estilo_tabela_pdf())
    elementos += [t2, Spacer(1, 16)]

    elementos.append(Paragraph("Inadimplentes", estilos["Heading2"]))
    dados3 = [["Cliente", "Telefone", "Valor em atraso"]]
    for c in Cliente.objects.filter(status="inadimplente"):
        dados3.append([c.nome, c.telefone, f"R$ {c.valor_mensal():.2f}"])
    t3 = Table(dados3, repeatRows=1)
    t3.setStyle(_estilo_tabela_pdf())
    elementos.append(t3)

    doc.build(elementos)
    return response


# ---------------------------------------------------------------------------
# RELATÓRIOS (somente admin)
# ---------------------------------------------------------------------------

MESES_PT = [
    "", "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
    "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro",
]


@user_passes_test(lambda u: u.is_authenticated and u.role == "admin")
def relatorio_tecnicos(request):
    hoje = timezone.now()
    ano = int(request.GET.get("ano", hoje.year))
    mes = int(request.GET.get("mes", hoje.month))

    tecnicos = User.objects.filter(role="tecnico")
    dados = []
    for t in tecnicos:
        chamados_tecnico = Chamado.objects.filter(tecnico=t)
        chamados_mes = chamados_tecnico.filter(criado_em__year=ano, criado_em__month=mes)
        dados.append({
            "tecnico": t,
            "concluidos_mes": chamados_mes.filter(status="concluido").count(),
            "concluidos_total": chamados_tecnico.filter(status="concluido").count(),
            "em_andamento": chamados_tecnico.filter(status="andamento").count(),
            "cancelados": chamados_tecnico.filter(status="cancelado").count(),
        })

    meses = [(i, MESES_PT[i]) for i in range(1, 13)]
    anos = list(range(hoje.year - 2, hoje.year + 1))

    context = {
        "dados": dados, "mes_selecionado": mes, "ano_selecionado": ano,
        "meses": meses, "anos": anos, "nome_mes": MESES_PT[mes],
    }
    return render(request, "core/relatorio_tecnicos.html", context)


@user_passes_test(lambda u: u.is_authenticated and u.role == "admin")
def backup_view(request):
    return render(request, "core/backup.html", {"backups": listar_backups()})


@user_passes_test(lambda u: u.is_authenticated and u.role == "admin")
def backup_criar_agora(request):
    if request.method == "POST":
        destino = fazer_backup()
        if destino:
            messages.success(request, f"Backup criado agora: {destino.name}")
        else:
            messages.error(request, "Não foi possível criar o backup (db.sqlite3 não encontrado).")
    return redirect("backup_view")


@user_passes_test(lambda u: u.is_authenticated and u.role == "admin")
def backup_download(request, nome):
    if not re.match(r"^backup_[\d\-_]+\.sqlite3$", nome):
        raise Http404()
    caminho = BACKUPS_DIR / nome
    if not caminho.exists():
        raise Http404()
    return FileResponse(open(caminho, "rb"), as_attachment=True, filename=nome)


@user_passes_test(lambda u: u.is_authenticated and u.role == "admin")
def backup_restaurar(request, nome):
    if not re.match(r"^backup_[\d\-_]+\.sqlite3$", nome):
        raise Http404()
    caminho = BACKUPS_DIR / nome
    if not caminho.exists():
        raise Http404()

    if request.method == "POST":
        fazer_backup()  # guarda o estado atual antes de sobrescrever, por segurança
        connections.close_all()
        shutil.copy2(caminho, settings.BASE_DIR / "db.sqlite3")
        messages.success(
            request,
            f"Banco de dados restaurado a partir de \"{nome}\". O sistema agora está exatamente "
            f"como estava quando esse backup foi salvo.",
        )
        return redirect("backup_view")

    return render(request, "core/backup_confirmar_restaurar.html", {"nome": nome})


@user_passes_test(lambda u: u.is_authenticated and u.role == "admin")
def backup_upload_restaurar(request):
    if request.method == "POST":
        arquivo = request.FILES.get("arquivo_backup")
        if not arquivo:
            messages.error(request, "Escolha um arquivo de backup (.sqlite3) antes de enviar.")
            return redirect("backup_view")
        if not arquivo.name.lower().endswith((".sqlite3", ".db")):
            messages.error(request, "O arquivo precisa ser um backup .sqlite3 (o mesmo tipo que o sistema gera).")
            return redirect("backup_view")

        fazer_backup()  # guarda o estado atual antes de sobrescrever, por segurança

        BACKUPS_DIR.mkdir(exist_ok=True)
        nome_temp = f"enviado_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.sqlite3"
        destino_temp = BACKUPS_DIR / nome_temp
        with open(destino_temp, "wb") as f:
            for pedaco in arquivo.chunks():
                f.write(pedaco)

        connections.close_all()
        shutil.copy2(destino_temp, settings.BASE_DIR / "db.sqlite3")
        messages.success(request, "Backup enviado e restaurado com sucesso! O sistema já está igual ao arquivo que você subiu.")
        return redirect("backup_view")

    return redirect("backup_view")


@user_passes_test(lambda u: u.is_authenticated and u.role == "admin")
def relatorios_view(request):
    hoje = timezone.now().date()
    ano = int(request.GET.get("ano", hoje.year))
    mes = int(request.GET.get("mes", hoje.month))
    tecnico_filtro = request.GET.get("tecnico", "")
    operador_filtro = request.GET.get("operador", "")

    # --- 1. Técnicos: quantos chamados cada um atendeu ---
    tecnicos_qs = User.objects.filter(role="tecnico")
    if tecnico_filtro:
        tecnicos_qs = tecnicos_qs.filter(pk=tecnico_filtro)
    dados_tecnicos = []
    for t in tecnicos_qs:
        chamados_t = Chamado.objects.filter(tecnico=t)
        chamados_mes_t = chamados_t.filter(criado_em__year=ano, criado_em__month=mes)
        dados_tecnicos.append({
            "pessoa": t,
            "concluidos_mes": chamados_mes_t.filter(status="concluido").count(),
            "em_andamento": chamados_t.filter(status="andamento").count(),
            "cancelados": chamados_t.filter(status="cancelado").count(),
            "concluidos_total": chamados_t.filter(status="concluido").count(),
        })

    # --- 2. Operadores: quantos chamados cada um abriu ---
    operadores_qs = User.objects.filter(role__in=["operador", "admin"])
    if operador_filtro:
        operadores_qs = operadores_qs.filter(pk=operador_filtro)
    dados_operadores = []
    for o in operadores_qs:
        qtd = Chamado.objects.filter(aberto_por=o, criado_em__year=ano, criado_em__month=mes).count()
        dados_operadores.append({"pessoa": o, "chamados_abertos": qtd})

    # --- 3. Total de chamados do mês (por tipo, por status) e ranking histórico ---
    chamados_mes_qs = Chamado.objects.filter(criado_em__year=ano, criado_em__month=mes)
    total_chamados_mes = chamados_mes_qs.count()

    por_tipo = []
    for valor, rotulo in Chamado.TIPO_CHOICES:
        qtd = chamados_mes_qs.filter(tipo=valor).count()
        if qtd:
            por_tipo.append({"rotulo": rotulo, "qtd": qtd})

    por_status = []
    for valor, rotulo in Chamado.STATUS_CHOICES:
        por_status.append({"rotulo": rotulo, "qtd": chamados_mes_qs.filter(status=valor).count()})

    contagem_meses = {}
    for c in Chamado.objects.all():
        chave = (c.criado_em.year, c.criado_em.month)
        contagem_meses[chave] = contagem_meses.get(chave, 0) + 1
    ranking_meses = sorted(contagem_meses.items(), key=lambda x: x[1], reverse=True)[:6]
    ranking_meses = [{"mes_ano": f"{MESES_PT[k[1]]}/{k[0]}", "qtd": v} for k, v in ranking_meses]

    # --- 4. Clientes: novos (instalação), cancelados, retornos no mês ---
    novos_count = MovimentacaoReceita.objects.filter(
        tipo="novo_cliente", criado_em__year=ano, criado_em__month=mes
    ).count()
    cancelados_count = MovimentacaoReceita.objects.filter(
        tipo="cancelamento", criado_em__year=ano, criado_em__month=mes
    ).count()
    retornos_qs = Chamado.objects.filter(
        eh_retorno=True, criado_em__year=ano, criado_em__month=mes
    ).select_related("cliente", "tecnico_ultimo_atendimento")

    # --- 5. Financeiro / crescimento: saldo do mês e tendência dos últimos 12 meses ---
    movs_mes = MovimentacaoReceita.objects.filter(criado_em__year=ano, criado_em__month=mes)
    receita_nova = sum((m.valor_novo for m in movs_mes.filter(tipo="novo_cliente")), 0)
    receita_perdida = sum((m.valor_anterior for m in movs_mes.filter(tipo="cancelamento")), 0)
    ajuste_planos = (
        sum((m.diferenca() for m in movs_mes.filter(tipo="upgrade")), 0)
        + sum((m.diferenca() for m in movs_mes.filter(tipo="downgrade")), 0)
    )
    saldo_mes = receita_nova - receita_perdida + ajuste_planos
    esperado_mes = sum((c.valor_mensal() for c in Cliente.objects.exclude(status="inativo")), 0)
    percentual_saldo = (saldo_mes / esperado_mes * 100) if esperado_mes else 0

    tendencia = []
    base_mes = date(ano, mes, 1)
    for i in range(11, -1, -1):
        ano_i = base_mes.year + ((base_mes.month - i - 1) // 12)
        mes_i = ((base_mes.month - i - 1) % 12) + 1
        movs_i = MovimentacaoReceita.objects.filter(criado_em__year=ano_i, criado_em__month=mes_i)
        rn = sum((m.valor_novo for m in movs_i.filter(tipo="novo_cliente")), 0)
        rp = sum((m.valor_anterior for m in movs_i.filter(tipo="cancelamento")), 0)
        aj = (
            sum((m.diferenca() for m in movs_i.filter(tipo="upgrade")), 0)
            + sum((m.diferenca() for m in movs_i.filter(tipo="downgrade")), 0)
        )
        tendencia.append({"mes_ano": f"{MESES_PT[mes_i][:3]}/{ano_i}", "saldo": rn - rp + aj})

    context = {
        "mes": mes, "ano": ano, "mes_ref": date(ano, mes, 1),
        "meses_opcoes": [(i, MESES_PT[i]) for i in range(1, 13)],
        "anos_opcoes": list(range(hoje.year - 2, hoje.year + 1)),
        "tecnicos_todos": User.objects.filter(role="tecnico"),
        "operadores_todos": User.objects.filter(role__in=["operador", "admin"]),
        "tecnico_filtro": tecnico_filtro,
        "operador_filtro": operador_filtro,
        "dados_tecnicos": dados_tecnicos,
        "dados_operadores": dados_operadores,
        "total_chamados_mes": total_chamados_mes,
        "por_tipo": por_tipo,
        "por_status": por_status,
        "ranking_meses": ranking_meses,
        "novos_count": novos_count,
        "cancelados_count": cancelados_count,
        "retornos_count": retornos_qs.count(),
        "retornos_lista": retornos_qs,
        "receita_nova": receita_nova,
        "receita_perdida": receita_perdida,
        "ajuste_planos": ajuste_planos,
        "saldo_mes": saldo_mes,
        "percentual_saldo": percentual_saldo,
        "tendencia": tendencia,
    }
    return render(request, "core/relatorios.html", context)


# ---------------------------------------------------------------------------
# EXPORTAÇÃO: EXCEL E PDF (admin + operador)
# ---------------------------------------------------------------------------

def _estilo_tabela_pdf():
    return TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563eb")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f1f5f9")]),
    ])


@somente_operacao
def cliente_export_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clientes"
    cabecalho = [
        "Nome", "CPF/CNPJ", "Data de Nascimento", "Telefone", "Rua", "Número", "Bairro", "Cidade",
        "Plano", "CTO", "Porta", "Login PPPoE", "Senha PPPoE",
        "Equipamento (de quem)", "Tipo de Equipamento", "Status", "Mensalidade",
    ]
    ws.append(cabecalho)
    for c in Cliente.objects.select_related("plano", "cto").all():
        ws.append([
            c.nome, c.cpf_cnpj,
            c.data_nascimento.strftime("%d/%m/%Y") if c.data_nascimento else "",
            c.telefone, c.logradouro, c.numero, c.bairro, c.cidade,
            str(c.plano) if c.plano else "", str(c.cto) if c.cto else "", c.porta,
            c.login_pppoe, c.senha_pppoe,
            c.get_propriedade_equipamento_display(), c.get_tipo_equipamento_display(),
            c.get_status_display(), float(c.valor_mensal()),
        ])
    for i in range(1, len(cabecalho) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 20

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="clientes.xlsx"'
    wb.save(response)
    return response


@somente_operacao
def cliente_export_pdf(request):
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="clientes.pdf"'
    doc = SimpleDocTemplate(response, pagesize=landscape(A4), topMargin=1.5 * cm, bottomMargin=1.5 * cm)
    estilos = getSampleStyleSheet()
    elementos = [Paragraph("Relatório de Clientes - CTO Manager Pro", estilos["Title"]), Spacer(1, 12)]

    dados = [["Nome", "Nascimento", "Endereço", "Login / Senha", "Equipamento", "CTO/Porta", "Status", "Mensalidade"]]
    for c in Cliente.objects.select_related("plano", "cto").all():
        endereco = f"{c.logradouro or '-'}, {c.numero or 's/n'} - {c.bairro or '-'}"
        login_senha = f"{c.login_pppoe or '-'} / {c.senha_pppoe or '-'}"
        equipamento = f"{c.get_propriedade_equipamento_display()} ({c.get_tipo_equipamento_display()})"
        dados.append([
            c.nome,
            c.data_nascimento.strftime("%d/%m/%Y") if c.data_nascimento else "-",
            endereco, login_senha, equipamento,
            f"{c.cto or '-'} / {c.porta or '-'}", c.get_status_display(), f"R$ {c.valor_mensal():.2f}",
        ])
    tabela = Table(dados, repeatRows=1)
    tabela.setStyle(_estilo_tabela_pdf())
    elementos.append(tabela)
    doc.build(elementos)
    return response


@somente_operacao
def cto_export_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CTOs"
    cabecalho = ["Código", "Bairro", "Endereço", "Capacidade", "Portas Ocupadas", "Portas Livres", "% Ocupação"]
    ws.append(cabecalho)
    for cto in CTO.objects.all():
        ws.append([
            cto.codigo, cto.bairro, cto.endereco, cto.capacidade,
            cto.portas_ocupadas(), cto.portas_livres(), cto.percentual_ocupacao(),
        ])
    for i in range(1, len(cabecalho) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 20

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="ctos.xlsx"'
    wb.save(response)
    return response


@somente_operacao
def cto_export_pdf(request):
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="ctos.pdf"'
    doc = SimpleDocTemplate(response, pagesize=landscape(A4), topMargin=1.5 * cm, bottomMargin=1.5 * cm)
    estilos = getSampleStyleSheet()
    elementos = [Paragraph("Relatório de CTOs - CTO Manager Pro", estilos["Title"]), Spacer(1, 12)]

    dados = [["Código", "Bairro", "Capacidade", "Ocupadas", "Livres", "% Ocupação"]]
    for cto in CTO.objects.all():
        dados.append([
            cto.codigo, cto.bairro, cto.capacidade,
            cto.portas_ocupadas(), cto.portas_livres(), f"{cto.percentual_ocupacao()}%",
        ])
    tabela = Table(dados, repeatRows=1)
    tabela.setStyle(_estilo_tabela_pdf())
    elementos.append(tabela)
    doc.build(elementos)
    return response


# ---------------------------------------------------------------------------
# ESTOQUE DE MATERIAL (admin + operador)
# ---------------------------------------------------------------------------
class MaterialListView(SomenteOperacaoMixin, ListView):
    model = Material
    template_name = "core/material_list.html"
    context_object_name = "materiais"

    def get_queryset(self):
        return Material.objects.filter(ativo=True)


class MaterialCreateView(SomenteOperacaoMixin, CreateView):
    model = Material
    form_class = MaterialForm
    template_name = "core/material_form.html"
    success_url = reverse_lazy("material_list")

    def form_valid(self, form):
        messages.success(self.request, "Material cadastrado no estoque!")
        return super().form_valid(form)


class MaterialUpdateView(SomenteOperacaoMixin, UpdateView):
    model = Material
    form_class = MaterialForm
    template_name = "core/material_form.html"
    success_url = reverse_lazy("material_list")

    def form_valid(self, form):
        messages.success(self.request, "Material atualizado!")
        return super().form_valid(form)


@somente_operacao
def estoque_entrada(request):
    if request.method == "POST":
        form = EntradaEstoqueForm(request.POST)
        if form.is_valid():
            material = form.cleaned_data["material"]
            quantidade = form.cleaned_data["quantidade"]
            MovimentacaoEstoque.objects.create(
                material=material, tipo="entrada", quantidade=quantidade,
                observacao=form.cleaned_data["observacao"], registrado_por=request.user,
            )
            messages.success(
                request,
                f"Entrada registrada: +{quantidade} {material.get_unidade_medida_display()} de "
                f"\"{material.nome}\". Saldo atual: {material.saldo_atual()}.",
            )
            return redirect("material_list")
    else:
        form = EntradaEstoqueForm()
    return render(request, "core/estoque_entrada_form.html", {"form": form})


@somente_operacao
def estoque_saida(request):
    if request.method == "POST":
        form = SaidaEstoqueForm(request.POST)
        if form.is_valid():
            material = form.cleaned_data["material"]
            quantidade = form.cleaned_data["quantidade"]
            MovimentacaoEstoque.objects.create(
                material=material, tipo="saida", quantidade=quantidade,
                tecnico=form.cleaned_data.get("tecnico"),
                observacao=form.cleaned_data["observacao"], registrado_por=request.user,
            )
            tecnico = form.cleaned_data.get("tecnico")
            destino = f" para {tecnico.get_full_name() or tecnico.username}" if tecnico else ""
            messages.success(
                request,
                f"Retirada registrada: -{quantidade} {material.get_unidade_medida_display()} de "
                f"\"{material.nome}\"{destino}. Saldo atual: {material.saldo_atual()}.",
            )
            return redirect("material_list")
    else:
        form = SaidaEstoqueForm()
    return render(request, "core/estoque_saida_form.html", {"form": form})


@somente_operacao
def estoque_historico(request):
    movimentacoes = MovimentacaoEstoque.objects.select_related("material", "tecnico", "chamado", "registrado_por")
    material_id = request.GET.get("material")
    if material_id:
        movimentacoes = movimentacoes.filter(material_id=material_id)
    context = {
        "movimentacoes": movimentacoes,
        "materiais": Material.objects.filter(ativo=True),
        "material_filtro": material_id or "",
    }
    return render(request, "core/estoque_historico.html", context)


# ---------------------------------------------------------------------------
# PONTO (Operador e Técnico batem o próprio ponto; Admin monitora e libera)
# ---------------------------------------------------------------------------
def _somente_operador_ou_tecnico(user):
    return user.is_authenticated and user.role in ("operador", "tecnico")


@user_passes_test(_somente_operador_ou_tecnico)
def ponto_bater(request):
    hoje = timezone.localdate()
    proximo = proximo_tipo_ponto(request.user, hoje)

    if request.method == "POST" and proximo:
        eh_chamada = False
        if proximo == "entrada":
            eh_chamada = LiberacaoExtraPonto.objects.filter(usuario=request.user, data=hoje, usada=False).exists()
            LiberacaoExtraPonto.objects.filter(usuario=request.user, data=hoje, usada=False).update(usada=True)
        RegistroPonto.objects.create(
            usuario=request.user, tipo=proximo, data_hora=timezone.now(), eh_chamada_de_volta=eh_chamada
        )
        messages.success(request, f"Ponto registrado: {dict(RegistroPonto.TIPO_CHOICES)[proximo]} às {timezone.localtime().strftime('%H:%M')}.")
        return redirect("ponto_bater")

    resumo = resumo_ponto_dia(request.user, hoje)
    jornada, _ = JornadaTrabalho.objects.get_or_create(usuario=request.user)
    chamado_de_volta = bool(proximo == "entrada" and resumo["todos_registros"])
    ultimo_registro = resumo["todos_registros"][-1] if resumo["todos_registros"] else None
    em_chamada_de_volta = bool(
        proximo == "saida" and ultimo_registro and ultimo_registro.tipo == "entrada" and ultimo_registro.eh_chamada_de_volta
    )
    context = {
        "resumo": resumo,
        "proximo": proximo,
        "proximo_label": dict(RegistroPonto.TIPO_CHOICES).get(proximo),
        "jornada": jornada,
        "hoje": hoje,
        "chamado_de_volta": chamado_de_volta,
        "em_chamada_de_volta": em_chamada_de_volta,
    }
    return render(request, "core/ponto_bater.html", context)


@user_passes_test(_somente_operador_ou_tecnico)
def ponto_meu_historico(request):
    hoje = timezone.localdate()
    ano = int(request.GET.get("ano", hoje.year))
    mes = int(request.GET.get("mes", hoje.month))

    primeiro_dia = date(ano, mes, 1)
    ultimo_dia_mes = (date(ano, mes + 1, 1) - timedelta(days=1)) if mes < 12 else date(ano, 12, 31)
    fim = min(ultimo_dia_mes, hoje) if (ano, mes) == (hoje.year, hoje.month) else ultimo_dia_mes

    dias = []
    data = primeiro_dia
    while data <= fim:
        resumo = resumo_ponto_dia(request.user, data)
        dias.append(resumo)
        data += timedelta(days=1)

    total_trabalhadas = sum(d["horas_trabalhadas"] for d in dias)
    total_esperadas = sum(d["horas_esperadas"] for d in dias)
    context = {
        "dias": dias,
        "total_trabalhadas": round(total_trabalhadas, 2),
        "total_esperadas": round(total_esperadas, 2),
        "total_diferenca": round(total_trabalhadas - total_esperadas, 2),
        "mes": mes, "ano": ano,
        "meses_opcoes": [(i, MESES_PT[i]) for i in range(1, 13)],
        "anos_opcoes": list(range(hoje.year - 1, hoje.year + 1)),
    }
    return render(request, "core/ponto_meu_historico.html", context)


@somente_admin
def ponto_admin_painel(request):
    hoje = timezone.localdate()
    funcionarios = User.objects.filter(role__in=["operador", "tecnico"], is_active=True).order_by("first_name", "username")
    dados = []
    for f in funcionarios:
        resumo = resumo_ponto_dia(f, hoje)
        liberacao_pendente = LiberacaoExtraPonto.objects.filter(usuario=f, data=hoje, usada=False).exists()
        dados.append({"usuario": f, "resumo": resumo, "liberacao_pendente": liberacao_pendente})
    context = {"dados": dados, "hoje": hoje}
    return render(request, "core/ponto_admin_painel.html", context)


@somente_admin
def ponto_liberar_mais_cedo(request):
    if request.method == "POST":
        form = PontoLiberarForm(request.POST)
        if form.is_valid():
            usuario = form.cleaned_data["usuario"]
            proximo = proximo_tipo_ponto(usuario, timezone.localdate())
            if not proximo:
                messages.warning(request, f"{usuario.get_full_name() or usuario.username} já bateu todos os pontos de hoje.")
                return redirect("ponto_admin_painel")
            RegistroPonto.objects.create(
                usuario=usuario, tipo=proximo, data_hora=timezone.now(),
                liberado_mais_cedo=True, autorizado_por=request.user,
                observacao=form.cleaned_data["observacao"],
            )
            messages.success(
                request,
                f"Registro liberado para {usuario.get_full_name() or usuario.username}: "
                f"{dict(RegistroPonto.TIPO_CHOICES)[proximo]} às {timezone.localtime().strftime('%H:%M')}.",
            )
            return redirect("ponto_admin_painel")
    else:
        form = PontoLiberarForm()
    return render(request, "core/ponto_liberar_form.html", {"form": form})


@somente_admin
def ponto_liberar_acesso_extra(request):
    if request.method == "POST":
        form = LiberacaoExtraForm(request.POST)
        if form.is_valid():
            usuario = form.cleaned_data["usuario"]
            LiberacaoExtraPonto.objects.create(
                usuario=usuario, data=timezone.localdate(),
                motivo=form.cleaned_data["motivo"], autorizado_por=request.user,
            )
            messages.success(
                request,
                f"Acesso liberado! Na próxima vez que {usuario.get_full_name() or usuario.username} "
                f"entrar no sistema, vai poder bater o ponto de entrada de novo.",
            )
            return redirect("ponto_admin_painel")
    else:
        form = LiberacaoExtraForm()
    return render(request, "core/ponto_liberar_acesso_extra_form.html", {"form": form})


@somente_admin
def abono_criar(request):
    if request.method == "POST":
        form = AbonoForm(request.POST, request.FILES)
        if form.is_valid():
            defaults = {"motivo": form.cleaned_data["motivo"], "registrado_por": request.user}
            if form.cleaned_data.get("anexo"):
                defaults["anexo"] = form.cleaned_data["anexo"]
            AbonoPonto.objects.update_or_create(
                usuario=form.cleaned_data["usuario"], data=form.cleaned_data["data"],
                defaults=defaults,
            )
            usuario = form.cleaned_data["usuario"]
            messages.success(
                request,
                f"Dia {form.cleaned_data['data'].strftime('%d/%m/%Y')} abonado pra "
                f"{usuario.get_full_name() or usuario.username}. Não vai contar como falta.",
            )
            return redirect("ponto_admin_painel")
    else:
        form = AbonoForm()
    return render(request, "core/abono_form.html", {"form": form})


@somente_admin
def jornada_editar(request, usuario_id):
    funcionario = get_object_or_404(User, pk=usuario_id, role__in=["operador", "tecnico"])
    jornada, _ = JornadaTrabalho.objects.get_or_create(usuario=funcionario)
    if request.method == "POST":
        form = JornadaForm(request.POST, instance=jornada)
        if form.is_valid():
            form.save()
            messages.success(request, f"Jornada de {funcionario.get_full_name() or funcionario.username} atualizada!")
            return redirect("ponto_admin_painel")
    else:
        form = JornadaForm(instance=jornada)
    return render(request, "core/jornada_form.html", {"form": form, "funcionario": funcionario})


@somente_admin
def ponto_relatorio(request):
    hoje = timezone.localdate()
    ano = int(request.GET.get("ano", hoje.year))
    mes = int(request.GET.get("mes", hoje.month))
    usuario_id = request.GET.get("usuario", "")

    funcionarios = User.objects.filter(role__in=["operador", "tecnico"], is_active=True).order_by("first_name", "username")
    alvo = funcionarios.filter(pk=usuario_id) if usuario_id else funcionarios

    primeiro_dia = date(ano, mes, 1)
    ultimo_dia_mes = (date(ano + (mes // 12), (mes % 12) + 1, 1) - timedelta(days=1)) if mes < 12 else date(ano, 12, 31)
    fim = min(ultimo_dia_mes, hoje) if (ano, mes) == (hoje.year, hoje.month) else ultimo_dia_mes

    resultados = []
    for f in alvo:
        dias = []
        data = primeiro_dia
        while data <= fim:
            resumo = resumo_ponto_dia(f, data)
            dias.append(resumo)
            data += timedelta(days=1)
        total_trabalhadas = sum(d["horas_trabalhadas"] for d in dias)
        total_esperadas = sum(d["horas_esperadas"] for d in dias)
        total_horas_extra = sum(d["diferenca"] for d in dias if d["diferenca"] > 0)
        total_horas_falta = sum(-d["diferenca"] for d in dias if d["diferenca"] < 0)
        abonos_mes = AbonoPonto.objects.filter(
            usuario=f, data__gte=primeiro_dia, data__lte=ultimo_dia_mes
        ).select_related("registrado_por").order_by("data")
        resultados.append({
            "usuario": f, "dias": dias,
            "total_trabalhadas": round(total_trabalhadas, 2),
            "total_esperadas": round(total_esperadas, 2),
            "total_diferenca": round(total_trabalhadas - total_esperadas, 2),
            "total_horas_extra": round(total_horas_extra, 2),
            "total_horas_falta": round(total_horas_falta, 2),
            "abonos_mes": abonos_mes,
        })

    context = {
        "resultados": resultados,
        "funcionarios": funcionarios,
        "usuario_filtro": usuario_id,
        "mes": mes, "ano": ano,
        "meses_opcoes": [(i, MESES_PT[i]) for i in range(1, 13)],
        "anos_opcoes": list(range(hoje.year - 1, hoje.year + 1)),
    }
    return render(request, "core/ponto_relatorio.html", context)


@somente_admin
def ponto_acumulado(request):
    """Soma, mês a mês, desde o primeiro registro de ponto da pessoa até hoje —
    pra mostrar quanta hora extra (ou falta) ela tem ACUMULADA no total, sem
    misturar com o mês atual (que fica só no Relatório mensal)."""
    hoje = timezone.localdate()
    usuario_id = request.GET.get("usuario", "")
    funcionarios = User.objects.filter(role__in=["operador", "tecnico"], is_active=True).order_by("first_name", "username")
    alvo = funcionarios.filter(pk=usuario_id) if usuario_id else funcionarios

    resultados = []
    for f in alvo:
        primeiro_registro = RegistroPonto.objects.filter(usuario=f).order_by("data_hora").first()
        if not primeiro_registro:
            continue
        cursor = primeiro_registro.data_hora.date().replace(day=1)

        meses = []
        while (cursor.year, cursor.month) <= (hoje.year, hoje.month):
            ano_m, mes_m = cursor.year, cursor.month
            primeiro_dia = date(ano_m, mes_m, 1)
            ultimo_dia_mes = (date(ano_m, mes_m + 1, 1) - timedelta(days=1)) if mes_m < 12 else date(ano_m, 12, 31)
            fim = min(ultimo_dia_mes, hoje) if (ano_m, mes_m) == (hoje.year, hoje.month) else ultimo_dia_mes

            total_trabalhadas = total_esperadas = total_extra = total_falta = 0.0
            data = primeiro_dia
            while data <= fim:
                resumo = resumo_ponto_dia(f, data)
                total_trabalhadas += resumo["horas_trabalhadas"]
                total_esperadas += resumo["horas_esperadas"]
                if resumo["diferenca"] > 0:
                    total_extra += resumo["diferenca"]
                else:
                    total_falta += -resumo["diferenca"]
                data += timedelta(days=1)

            meses.append({
                "mes": mes_m, "ano": ano_m, "mes_nome": MESES_PT[mes_m],
                "trabalhadas": round(total_trabalhadas, 2),
                "esperadas": round(total_esperadas, 2),
                "extra": round(total_extra, 2),
                "falta": round(total_falta, 2),
                "saldo": round(total_trabalhadas - total_esperadas, 2),
            })

            cursor = date(ano_m + 1, 1, 1) if mes_m == 12 else date(ano_m, mes_m + 1, 1)

        # Mês mais recente primeiro na tela
        meses.reverse()
        resultados.append({
            "usuario": f,
            "meses": meses,
            "acumulado_extra": round(sum(m["extra"] for m in meses), 2),
            "acumulado_falta": round(sum(m["falta"] for m in meses), 2),
            "acumulado_saldo": round(sum(m["saldo"] for m in meses), 2),
        })

    context = {
        "resultados": resultados,
        "funcionarios": funcionarios,
        "usuario_filtro": usuario_id,
    }
    return render(request, "core/ponto_acumulado.html", context)


@somente_admin
def ponto_relatorio_pdf(request):
    hoje = timezone.localdate()
    ano = int(request.GET.get("ano", hoje.year))
    mes = int(request.GET.get("mes", hoje.month))
    usuario_id = request.GET.get("usuario", "")

    funcionarios = User.objects.filter(role__in=["operador", "tecnico"], is_active=True).order_by("first_name", "username")
    alvo = funcionarios.filter(pk=usuario_id) if usuario_id else funcionarios

    primeiro_dia = date(ano, mes, 1)
    ultimo_dia_mes = (date(ano + (mes // 12), (mes % 12) + 1, 1) - timedelta(days=1)) if mes < 12 else date(ano, 12, 31)
    fim = min(ultimo_dia_mes, hoje) if (ano, mes) == (hoje.year, hoje.month) else ultimo_dia_mes

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="relatorio_ponto_{mes:02d}_{ano}.pdf"'
    doc = SimpleDocTemplate(response, pagesize=landscape(A4), topMargin=1.5 * cm, bottomMargin=1.5 * cm)
    estilos = getSampleStyleSheet()
    elementos = [Paragraph(f"Relatório de Ponto - {MESES_PT[mes]}/{ano}", estilos["Title"]), Spacer(1, 12)]

    def hhmm(reg):
        return timezone.localtime(reg.data_hora).strftime("%H:%M") if reg else "-"

    for f in alvo:
        dias = []
        data = primeiro_dia
        while data <= fim:
            resumo = resumo_ponto_dia(f, data)
            dias.append(resumo)
            data += timedelta(days=1)
        total_trabalhadas = sum(d["horas_trabalhadas"] for d in dias)
        total_esperadas = sum(d["horas_esperadas"] for d in dias)
        total_diferenca = total_trabalhadas - total_esperadas

        nome = f.get_full_name() or f.username
        elementos.append(Paragraph(
            f"{nome} ({f.get_role_display()}) — Trabalhadas: {total_trabalhadas:.2f}h · "
            f"Esperadas: {total_esperadas:.2f}h · Saldo: {total_diferenca:.2f}h",
            estilos["Heading3"],
        ))
        dados = [["Data", "Entrada", "S. Almoço", "V. Almoço", "Saída", "Extra (chamado)", "Trabalhadas", "Esperadas", "Diferença"]]
        for d in dias:
            marcador = ""
            if d["liberado_mais_cedo"]:
                marcador += " (Liberado)"
            if d["abono"]:
                marcador += " (Abonado)"
            dados.append([
                d["data"].strftime("%d/%m/%Y") + marcador,
                hhmm(d["registros"].get("entrada")),
                hhmm(d["registros"].get("saida_almoco")),
                hhmm(d["registros"].get("volta_almoco")),
                hhmm(d["registros"].get("saida")),
                f"{d['horas_chamada_de_volta']:.2f}h" if d["horas_chamada_de_volta"] > 0 else "-",
                f"{d['horas_trabalhadas']:.2f}h",
                f"{d['horas_esperadas']:.2f}h",
                f"{d['diferenca']:.2f}h",
            ])
        tabela = Table(dados, repeatRows=1)
        tabela.setStyle(_estilo_tabela_pdf())
        elementos += [tabela, Spacer(1, 16)]

    if not elementos[2:]:
        elementos.append(Paragraph("Nenhum funcionário encontrado.", estilos["Normal"]))

    doc.build(elementos)
    return response


# ---------------------------------------------------------------------------
# USUÁRIOS (somente admin) — criar funcionários e definir o perfil de cada um
# ---------------------------------------------------------------------------

class UsuarioListView(SomenteAdminMixin, ListView):
    model = User
    template_name = "core/usuario_list.html"
    context_object_name = "usuarios"

    def get_queryset(self):
        return User.objects.all().order_by("first_name", "username")


class UsuarioCreateView(SomenteAdminMixin, CreateView):
    model = User
    form_class = UsuarioCreateForm
    template_name = "core/usuario_form.html"
    success_url = reverse_lazy("usuario_list")

    def form_valid(self, form):
        messages.success(self.request, "Usuário cadastrado com sucesso!")
        return super().form_valid(form)


class UsuarioUpdateView(SomenteAdminMixin, UpdateView):
    model = User
    form_class = UsuarioUpdateForm
    template_name = "core/usuario_form.html"
    success_url = reverse_lazy("usuario_list")

    def form_valid(self, form):
        messages.success(self.request, "Usuário atualizado com sucesso!")
        return super().form_valid(form)


@user_passes_test(lambda u: u.is_authenticated and u.role == "admin")
def usuario_delete(request, pk):
    usuario = get_object_or_404(User, pk=pk)
    admins_ativos = User.objects.filter(role="admin", is_active=True).count()
    if usuario.role == "admin" and admins_ativos <= 1:
        messages.error(request, "Não é possível excluir o único administrador ativo do sistema.")
        return redirect("usuario_list")

    if request.method == "POST":
        nome = usuario.get_full_name() or usuario.username
        usuario.delete()
        messages.success(request, f"Usuário \"{nome}\" excluído do sistema.")
        return redirect("usuario_list")
    return render(request, "core/usuario_confirm_delete.html", {"usuario": usuario})


# ---------------------------------------------------------------------------
# GRUPOS E PERMISSÕES (somente admin)
# ---------------------------------------------------------------------------

MODELOS_PERMISSAO = ["cliente", "cto", "chamado", "plano", "contapagar", "user"]
ACOES_PERMISSAO = ["view", "add", "change", "delete"]


def _matriz_permissoes(grupo):
    ativas = set(grupo.permissions.filter(content_type__model__in=MODELOS_PERMISSAO).values_list("codename", flat=True))
    linhas = []
    for modelo in MODELOS_PERMISSAO:
        celulas = []
        for acao in ACOES_PERMISSAO:
            codename = f"{acao}_{modelo}"
            perm = Permission.objects.filter(codename=codename, content_type__model=modelo).first()
            celulas.append({
                "acao_label": ACAO_LABELS[acao],
                "perm_id": perm.id if perm else None,
                "ativo": codename in ativas,
            })
        linhas.append({"modelo_label": MODELO_LABELS.get(modelo, modelo.capitalize()), "celulas": celulas})
    return linhas


@user_passes_test(lambda u: u.is_authenticated and u.role == "admin")
def grupos_permissoes(request):
    if request.method == "POST":
        grupo_id = request.POST.get("grupo_id")
        grupo = get_object_or_404(Group, pk=grupo_id)
        ids_selecionados = request.POST.getlist("permissoes")
        grupo.permissions.set(ids_selecionados)
        messages.success(request, f'Permissões do grupo "{grupo.name}" atualizadas!')
        return redirect("grupos_permissoes")

    grupos = Group.objects.all().order_by("name")
    dados_grupos = [{"grupo": g, "linhas": _matriz_permissoes(g)} for g in grupos]
    return render(request, "core/grupos_permissoes.html", {"dados_grupos": dados_grupos})
